import re
from urllib.request import Request, urlopen
from io import BytesIO
from flask import Flask, jsonify, request
from urllib.error import HTTPError
import atexit

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger

# from auth import login_required, admin_required, SECRET_KEY
# from itsdangerous import (TimedJSONWebSignatureSerializer as Serializer)
from flask_cors import CORS
from mongoengine import connect
from openpyxl import load_workbook

from models import LGA, Year

app = Flask(__name__)
CORS(app, supports_credentials=True)
connect(
    host='mongodb://user:user@ds231360.mlab.com:31360/nsw_crime_data'
)


def get_column(ws, col_i):
    data = []
    for col in ws.iter_cols(min_col=col_i, max_col=col_i, min_row=8, max_row=69):
        for cell in col:
            if cell.value is not None:
                data.append(str(cell.value))
            else:
                data.append('')
    return data


def get_all_file_names():
    with urlopen(
            "http://www.bocsar.nsw.gov.au/Pages/bocsar_crime_stats/bocsar_lgaexceltables.aspx") as response:
        html = response.read().decode('utf-8')
        matched = re.findall("\"/Documents/RCS-Annual/(\w+\.xlsx)\">([^<>]+)<", html)
        f_names = [x[0].lower() for x in matched]
        lga_f_name_mapping = {x[1].lower().replace(' ', ''): x[0] for x in matched}
    return f_names, lga_f_name_mapping


def data2db(file_name, response):
    last_modified = response.getheader("last-modified")
    bytes_data = response.read()
    wb = load_workbook(filename=BytesIO(bytes_data))
    ws = wb[wb.sheetnames[0]]
    # lga_name = " ".join(ws.cell(row=5, column=1).value.split()[:-3])
    number_of_year = (ws.max_column - 5) // 2
    years = []
    if ws.cell(row=1, column=3).value is None:
        t = 0
        for i in range(number_of_year):
            str1 = ws.cell(row=6, column=3 + i * 2).value
            year = ''.join(list(filter(str.isdigit, str1)))
            rate = get_column(ws, 4 + i * 2)
            s = None
            for x in rate:
                if x != 'nc':
                    if s is None:
                        s = float(x)
                    else:
                        s += float(x)
            if s is not None:
                total_rate = s
            else:
                total_rate = 'Null'
                number_of_year -= 1
            if total_rate != 'Null':
                t += total_rate
            years.append(Year(year=str(year),crime_rate=str(total_rate)))
        if number_of_year != 0:
            average = t / number_of_year
        else:
            average = 'Null'
    else:
        t = 0
        for i in range(number_of_year):
            end_year = int(ws.cell(row=1, column=3).value)
            year = str(end_year - 4 + i)
            rate = get_column(ws, 4 + i * 2)
            s = None
            for x in rate:
                if x != 'nc':
                    if s is None:
                        s = float(x)
                    else:
                        s += float(x)
            if s is not None:
                total_rate = s
            else:
                total_rate = 'Null'
                number_of_year -= 1
            if total_rate != 'Null':
                t += total_rate
            years.append(Year(year=str(year), crime_rate=str(total_rate)))
        if number_of_year != 0:
            average = t / number_of_year
        else:
            average = 'Null'
    crime_data = LGA(file_name, years, str(average), last_modified)
    crime_data.save()


file_names, lga_file_name_mapping = get_all_file_names()


# for file_name in file_names:
#     print(file_name)
#     data2db(file_name)

@app.route("/nsw_crime_data", methods=["GET"])
# @login_required
def get_collections():
    lga_qs = LGA.objects()
    entries = []
    for lga in lga_qs:
        year_data = dict()
        for dt in lga.year_data:
            year_data[str(dt['year'])]=dt['crime_rate']
        entries.append({'id': '{}/{}'.format(request.base_url, lga.file_name[:-8]),
                        'updated': lga.last_modified,
                        'lga_name': lga.file_name[:-8],
                        'year_data': year_data,
                        'average': lga.average
                        })
    return jsonify(title='NSW Crime Statistics',
                   id=request.base_url,
                   entry=entries), 200


@app.route("/nsw_crime_data/<rid>", methods=["DELETE"])
# @admin_required
def delete_by_id(rid):
    lga_qs = LGA.objects(file_name=rid + 'lga.xlsx')
    if lga_qs.count() != 0:
        lga = lga_qs[0]
        lga.delete()
        return 'Delete Success.', 200
    else:
        return "Id not found.", 404


@app.route("/nsw_crime_data/<rid>", methods=["GET"])
# @login_required
def get_by_id(rid):
    lga_qs = LGA.objects(file_name=rid + 'lga.xlsx')
    if lga_qs.count() != 0:
        lga = lga_qs[0]
        year_data=dict()
        for dt in lga.year_data:
            year_data[str(dt['year'])]=dt['crime_rate']
        xml_dict = {
            'id': request.base_url,
            'updated': lga.last_modified,
            'lga_name': lga.file_name[:-8],
            'year_data': year_data,
            'average': lga.average
        }
        return jsonify(xml_dict), 200
    else:
        return "Id not found.", 404


def update_db():
    print("scheduled db update")
    for file_name in file_names:
        lga_qs = LGA.objects(file_name=file_name)
        if lga_qs.count() != 0:
            lga = lga_qs[0]
            r = Request('http://www.bocsar.nsw.gov.au/Documents/RCS-Annual/{}'.format(file_name))
            r.add_header('If-modified-since', lga.last_modified)
            try:
                with urlopen(r) as response:
                    print('updating ' + file_name)
                    data2db(file_name, response)
            except HTTPError:
                continue
        else:
            with urlopen(
                    'http://www.bocsar.nsw.gov.au/Documents/RCS-Annual/{}'.format(file_name)) as response:
                print('updating ' + file_name)
                data2db(file_name, response)


if __name__ == "__main__":
    # update_db()
    scheduler = BackgroundScheduler()
    scheduler.start()
    scheduler.add_job(
        func=update_db,
        trigger=IntervalTrigger(hours=1),
        id='updating_job',
        name='Update database every day',
        replace_existing=True)
    # Shut down the scheduler when exiting the app
    atexit.register(lambda: scheduler.shutdown())

    app.run(port=50102)